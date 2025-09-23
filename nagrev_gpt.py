from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
import os
from pathlib import Path
import traceback
import time
from openpyxl.drawing.image import Image
from openpyxl.worksheet.header_footer import HeaderFooter
import customtkinter as ctk
from tkinter import filedialog, messagebox, ttk
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.worksheet.pagebreak import RowBreak  # разрыв страниц
import sqlite3

# Добавьте эти строки
films_df = None
all_sidak = []
# Глобальная переменная для имени протокола (для использования при сохранении)
protocol_title = ""

# ============================
# УТИЛИТЫ ДЛЯ ОФОРМЛЕНИЯ
# ============================

def create_cell(ws, cell, value, bold=False, font_size=11, horizontal='left', vertical='center', wrap_text=False):
    """Вспомогательная функция для создания ячеек с оформлением"""
    ws[cell] = value
    ws[cell].font = Font(size=font_size, bold=bold)
    ws[cell].alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)


def setup_print_settings(ws):
    """Настраивает параметры печати"""
    try:
        # Установка области печати
        ws.print_area = f'A1:D{ws.max_row}'

        # Настройка страницы
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1

        # Поля
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3

        print("✅ Настройки печати применены")
    except Exception as e:
        print(f"⚠️ Ошибка при настройке печати: {e}")


# ============================
# ДОБАВЛЕНИЕ ИЗОБРАЖЕНИЙ (оригинальное диалоговое окно) 
# ============================

def add_images_to_test_table(ws, films_data):
    """
    Открывает диалоговое окно для добавления картинок ко всем пленкам по очереди
    :param ws: рабочий лист
    :param films_data: список данных о пленках
    """
    def add_image_for_selected_film():
        nonlocal current_film_index

        if current_film_index >= len(films_data):
            messagebox.showinfo("Завершено", "Все картинки добавлены!")
            root.destroy()
            return

        film = films_data[current_film_index]
        film_label.configure(text=f"Пленка: {film['sidak_num']}")

        # Находим строку с текущей пленкой в таблице испытаний
        target_row = None
        for row in range(1, ws.max_row + 1):
            cell_value = ws[f'A{row}'].value
            # Более точный поиск - ищем точное совпадение с номером пленки
            if cell_value and film['sidak_num'] in str(cell_value) and "09:30-17:30" in str(cell_value):
                target_row = row
                break

        if not target_row:
            print(f"⚠️ Не найдена строка для пленки {film['sidak_num']}")
            current_film_index += 1
            add_image_for_selected_film()
            return

        # Выбираем файл с картинкой
        file_path = filedialog.askopenfilename(
            title=f"Выберите изображение для {film['sidak_num']}",
            filetypes=[("Image files", "*.jpg *.jpeg *.png *.bmp *.gif")]
        )

        if file_path:
            try:
                # Открываем изображение для получения оригинальных размеров
                from PIL import Image as PILImage
                pil_img = PILImage.open(file_path)

                # Целевые максимальные размеры
                max_width = 250
                max_height = 250

                # Получаем оригинальные размеры
                orig_width, orig_height = pil_img.size

                # Вычисляем коэффициенты масштабирования
                width_ratio = max_width / orig_width
                height_ratio = max_height / orig_height

                # Выбираем минимальный коэффициент для сохранения пропорций
                scale_ratio = min(width_ratio, height_ratio)

                # Вычисляем новые размеры
                new_width = int(orig_width * scale_ratio)
                new_height = int(orig_height * scale_ratio)

                # Добавляем картинку в ячейку C (столбец 3)
                img = Image(file_path)

                # Устанавливаем пропорциональные размеры
                img.width = new_width
                img.height = new_height

                # Добавляем картинку в ячейку
                cell_address = f"C{target_row}"
                ws.add_image(img, cell_address)

                print(f"✅ Картинка добавлена для {film['sidak_num']} в ячейку {cell_address}")
                print(f"   Размеры: {new_width}x{new_height} (оригинал: {orig_width}x{orig_height})")

            except Exception as e:
                print(f"⚠️ Ошибка при добавлении картинки: {e}")

        # Переходим к следующей пленке
        current_film_index += 1
        if current_film_index < len(films_data):
            add_image_for_selected_film()
        else:
            messagebox.showinfo("Завершено", "Все картинки добавлены!")
            root.destroy()

    def skip_film():
        nonlocal current_film_index
        current_film_index += 1
        if current_film_index < len(films_data):
            add_image_for_selected_film()
        else:
            messagebox.showinfo("Завершено", "Обработка завершена!")
            root.destroy()

    def skip_all():
        messagebox.showinfo("Завершено", "Добавление картинок пропущено!")
        root.destroy()

    # Создаем GUI окно
    root = ctk.CTkToplevel()
    root.title("Добавление изображений к пленкам")
    root.geometry("500x200")

    current_film_index = 0

    # Информация о процессе
    ctk.CTkLabel(root, text="Добавление изображений для каждой пленки",
             font=("Arial", 12, "bold")).pack(pady=10)

    film_label = ctk.CTkLabel(root, text="", font=("Arial", 10))
    film_label.pack(pady=5)

    # Кнопки
    btn_frame = ctk.CTkFrame(root)
    btn_frame.pack(pady=20)

    add_btn = ctk.CTkButton(
        btn_frame,
        text="Добавить картинку",
        command=add_image_for_selected_film,
        fg_color="green",
        hover_color="darkgreen"
    )
    add_btn.pack(side=ctk.LEFT, padx=5)

    skip_btn = ctk.CTkButton(
        btn_frame,
        text="Пропустить эту",
        command=skip_film,
        fg_color="yellow",
        hover_color="darkyellow",
        text_color="black"
    )
    skip_btn.pack(side=ctk.LEFT, padx=5)

    skip_all_btn = ctk.CTkButton(
        btn_frame,
        text="Пропустить все",
        command=skip_all,
        fg_color="red",
        hover_color="darkred"
    )
    skip_all_btn.pack(side=ctk.LEFT, padx=5)

    # Запускаем процесс для первой пленки
    if films_data:
        film_label.configure(text=f"Пленка: {films_data[0]['sidak_num']}")
    else:
        messagebox.showinfo("Информация", "Нет данных о пленках")
        root.destroy()
        return

    root.mainloop()


# ============================
# ПОСТРОЕНИЕ ПРОТОКОЛА (в памяти) 
# ============================

def build_protocol(ws, films_data, test_type="Нагрев", author="", participant=""):
    """Заполняет переданный лист протоколом. НИЧЕГО не сохраняет на диск."""
    # Словарь для преобразования баллов в текст
    score_to_text = {
        5: "Никаких изменений",
        4: "Малозаметное стягивание пленки в обернутых пленках изделиях;",
        3: "Небольшой зазор",
        2: "Большой зазор/частичное отхождение кромки",
        1: "Кромка отошла на протяжении 1/3 длины"
    }

    # Динамический расчет строк
    current_row = 1

    # ШАПКА ПРОТОКОЛА
    create_cell(ws, f'D{current_row}', 'УТВЕРЖДАЮ:', bold=True, horizontal='right')
    create_cell(ws, f'D{current_row+1}', 'Генеральный директор', horizontal='right')
    create_cell(ws, f'D{current_row+2}', 'ООО «Сидак СП»', horizontal='right')
    create_cell(ws, f'D{current_row+3}', '_____________ Репичев Д.А.', horizontal='right')
    create_cell(ws, f'D{current_row+4}', '«____» ______________2025г.', horizontal='right')

    current_row += 6

    # НАЗВАНИЕ ПРОТОКОЛА
    film_numbers = ', '.join([film['sidak_num'] for film in films_data])
    current_date = datetime.now().strftime('%d.%m.%Yг.')
    protocol_title_local = f'Протокол испытаний на {test_type.lower()} фасада в пленках {film_numbers} от {current_date}'

    create_cell(ws, f'A{current_row}', protocol_title_local, bold=True, font_size=12, horizontal='center', wrap_text=True)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws.row_dimensions[current_row].height = 30
    current_row += 2

    # 1. ЦЕЛЬ ПРОВЕДЕНИЯ ИСПЫТАНИЯ
    create_cell(ws, f'A{current_row}', '1. Цель проведения испытания: оценить прочность покрытия при нагреве фасада в материалах:', bold=False, wrap_text=True)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # Заголовок таблицы материалов
    create_cell(ws, f'A{current_row}', 'Тип материала', bold=True, horizontal='center')
    create_cell(ws, f'B{current_row}', 'Номер Sidak', bold=True, horizontal='center')
    create_cell(ws, f'C{current_row}', 'Наименование у поставщика', bold=True, horizontal='center')
    create_cell(ws, f'D{current_row}', 'Толщина', bold=True, horizontal='center')
    current_row += 1

    # Заполняем таблицу материалов
    for film in films_data:
        create_cell(ws, f'A{current_row}', 'Пленка', horizontal='center')
        create_cell(ws, f'B{current_row}', film.get('sidak_num', ''), horizontal='center')
        create_cell(ws, f'C{current_row}', film.get('supplier_name', ''), horizontal='center')
        create_cell(ws, f'D{current_row}', film.get('thickness', ''), horizontal='center')
        current_row += 1

    current_row += 2  # Отступ после таблицы

    # 2. ПРОДУКЦИЯ
    create_cell(ws, f'A{current_row}', '2. Продукция: Детали для изготовления мебели в пленке ПВХ/ПЭТ', bold=True)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 1

    # 3. НД НА ПРОДУКЦИЮ
    create_cell(ws, f'A{current_row}', '3. НД на продукцию: ТУ 31.09.10-012-46275274-2022 Детали для изготовления мебели в пленке ПВХ/ПЭТ', bold=True, wrap_text=True)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # 4. МЕСТО ПРОВЕДЕНИЯ
    create_cell(ws, f'A{current_row}', '4. Место проведения замеров: ООО «Сидак-СП», Россия, Ленинградская обл., Гатчинский р-н, г.п. Сиверский, ул. Заводская 9, корп.2.', bold=True, wrap_text=True)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # 5. ПРИБОР
    create_cell(ws, f'A{current_row}', '5. Прибор для измерения: Сушильный шкаф ЕС-4610.', bold=True)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 1

    # 6. УСЛОВИЯ
    create_cell(ws, f'A{current_row}', '6. Условия проведения испытаний: температура окружающего воздуха – 20-25°С, влажность не измерялась.', bold=True, wrap_text=True)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # 7. ПРИСУТСТВОВАЛИ
    create_cell(ws, f'A{current_row}', '7. Присутствовали:', bold=True)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 1
    create_cell(ws, f'A{current_row}', f'{author}/{participant}', wrap_text=True)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # 8. МЕТОД ОПРЕДЕЛЕНИЯ
    create_cell(ws, f'A{current_row}', '8. Метод определения:  ГОСТ 32289-2013 Приложение Г. Проведение испытаний: Образцы помещают в при температуре (20 + 5) °С и осматривают декоративную поверхность невооруженным глазом в целях обнаружения трещин на поверхности. Осмотр образцов производят под углом 20° - 30° к плоскости поверхности с расстояния 250 мм.', bold=False, wrap_text=True)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws.row_dimensions[current_row].height = 60
    current_row += 2

    # ТАБЛИЦА №2 - МАТЕРИАЛЫ ДЛЯ ИЗГОТОВЛЕНИЯ ОБРАЗЦОВ
    create_cell(ws, f'A{current_row}', 'Таблица №2 Материалы для изготовления образцов', bold=True)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 1

    create_cell(ws, f'A{current_row}', '№ образца', bold=True, horizontal='center')
    create_cell(ws, f'B{current_row}', 'МДФ', bold=True, horizontal='center')
    create_cell(ws, f'C{current_row}', 'Клей', bold=True, horizontal='center')
    create_cell(ws, f'D{current_row}', 'Пленка', bold=True, horizontal='center')
    current_row += 1

    # Заполняем таблицу образцов (используем выбранные материал и клей, если переданы)
    for i, film in enumerate(films_data, 1):
        create_cell(ws, f'A{current_row}', i, horizontal='center')
        create_cell(ws, f'B{current_row}', film.get('material', 'МДФ белый 19 мм  Кроношпан'), horizontal='center', wrap_text=True)
        create_cell(ws, f'C{current_row}', film.get('glue', 'Клей Perfotak 154/3'), horizontal='center', wrap_text=True)
        create_cell(ws, f'D{current_row}', film.get('sidak_num', ''), horizontal='center', wrap_text=True)
        ws.row_dimensions[current_row].height = 30
        current_row += 1

    current_row += 2  # Отступ после таблицы

    # ТАБЛИЦА №3 - ИСПЫТАНИЯ ОБРАЗЦОВ
    create_cell(ws, f'A{current_row}', 'Таблица №3 Испытания образцов', bold=True)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 1

    create_cell(ws, f'A{current_row}', 'Время нагрева', bold=True, horizontal='center')
    create_cell(ws, f'B{current_row}', 't° C', bold=True, horizontal='center')
    create_cell(ws, f'C{current_row}', 'Состояние изделия', bold=True, horizontal='center')
    create_cell(ws, f'D{current_row}', 'Результат нагрева', bold=True, horizontal='center', wrap_text=True)
    current_row += 1

    # Заполняем таблицу испытаний
    for film in films_data:
        create_cell(ws, f'A{current_row}', f"{film.get('sidak_num', '')} 09:30-17:30", horizontal='center', wrap_text=True)
        create_cell(ws, f'B{current_row}', 70, horizontal='center', wrap_text=True)
        create_cell(ws, f'C{current_row}', '')
        create_cell(ws, f'D{current_row}', score_to_text.get(int(film.get('score', 0)), ''), horizontal='center', wrap_text=True)
        ws.row_dimensions[current_row].height = 200
        current_row += 1

    current_row += 2  # Отступ после таблицы

    # КРИТЕРИИ ОЦЕНКИ
    create_cell(ws, f'A{current_row}', 'Критерии для оценки результатов', bold=True)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 1

    criteria_texts = [
        '5 баллов - Никаких видимых изменений на кромке или клеевом соединении;',
        '4 балла - Незначительное изменение в окраске. Малозаметное стягивание пленки в обернутых пленкам изделиях;',
        '3 балла - Небольшой зазор на задней стороне панели или тенденция к формированию зазора в видимых точках в области кромки или угла;',
        '2 балла - Большой зазор в видимых точках клеевого соединения. Полностью или частично отхождение кромки по линии кромок или в области угла. Значительные изгибы кромок и значительное изменение цвета.Частичное прилипание поверхности кромки к термальному элементу;',
        '1 балл - Полностью или частично разрушенные клеевые соединения. Полностью или частично отошедшая кромка на протяжении 1/3 всей длины или более. Сильные изгибы кромок и необратимое изменение цвета.',
    ]

    for text in criteria_texts:
        create_cell(ws, f'A{current_row}', text, wrap_text=True)
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws.row_dimensions[current_row].height = 45
        current_row += 1

    current_row += 1  # Отступ после критериев

    # ВЫВОДЫ
    create_cell(ws, f'A{current_row}', 'Выводы:', bold=True)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 1

    create_cell(ws, f'A{current_row}', 'Материал', bold=True, horizontal='center')
    create_cell(ws, f'B{current_row}', 'Толщина', bold=True, horizontal='center')
    create_cell(ws, f'C{current_row}', 'Результат', bold=True, horizontal='center')
    create_cell(ws, f'D{current_row}', 'Оценка', bold=True, horizontal='center')
    current_row += 1

    # Заполняем таблицу выводов
    for film in films_data:
        score = int(film.get('score', 0))
        ball_text = "баллов"
        if score == 1:
            ball_text = "балл"
        elif 2 <= score <= 4:
            ball_text = "балла"

        create_cell(ws, f'A{current_row}', f"Пленка {film.get('sidak_num', '')}")
        create_cell(ws, f'B{current_row}', film.get('thickness', ''))
        # Результат (текст по баллам)
        create_cell(ws, f'C{current_row}', score_to_text.get(score, ''))
        create_cell(ws, f'D{current_row}', f"{score} {ball_text}", horizontal='center')
        current_row += 1

    current_row += 2  # Отступ после выводов

    # ПОДПИСИ
    create_cell(ws, f'A{current_row}', 'Составил:', bold=True)
    current_row += 1
    create_cell(ws, f'A{current_row}', f'{author}')
    ws.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 2

    create_cell(ws, f'A{current_row}', 'Участники испытания:', bold=True)
    current_row += 1
    create_cell(ws, f'A{current_row}', f'{author} / {participant}')
    ws.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 3

    # ФУТЕР
    create_cell(ws, f'A{current_row}', 'Внимание! Данная информация является интеллектуальной собственностью ООО «Сидак-СП»', bold=True, font_size=10)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 1
    create_cell(ws, f'A{current_row}', 'и предназначена для внутреннего пользования. Распространение без разрешения', font_size=10)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 1
    create_cell(ws, f'A{current_row}', 'правообладателя запрещается.', font_size=10)
    ws.merge_cells(f'A{current_row}:D{current_row}')

    # Ширины колонок и печать
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 25

    setup_print_settings(ws)
    # возвращаем title (строка с переносом \n)
    return protocol_title_local


def create_protocol(films_data, output_filename):
    """Создаёт новую книгу, заполняет протокол и СРАЗУ сохраняет на диск (совместимость со старым поведением)."""
    global protocol_title
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"
    # Сохраняем title в глобальной переменной
    protocol_title = build_protocol(ws, films_data)
    wb.save(output_filename)
    print(f"✅ Протокол создан: {output_filename}")
    return wb, ws


# ============================
# ГЛАВНАЯ ФОРМА ВВОДА
# ============================

def show_input_form():
    global films_df, all_sidak
    films_data = []
    wb = None
    ws = None

    # --- состояния для панели изображений ---
    current_film_index = 0

    def validate_and_collect():
        sidak = combo_sidak.get().strip()
        supplier = entry_supplier.get().strip()
        material = combo_material.get().strip()
        glue = combo_glue.get().strip()
        thick_raw = entry_thickness.get().strip().replace(',', '.')
        score_str = combo_score.get().strip()

        # Проверка, что выбранный Sidak существует в базе
        if sidak not in all_sidak:
            messagebox.showwarning("Ошибка", "Выбранный номер Sidak не существует в базе данных.")
            return None

        if not sidak:
            messagebox.showwarning("Проверка", "Укажите Обозначение Sidak")
            return None
        if not supplier:
            messagebox.showwarning("Проверка", "Укажите Наименование у поставщика")
            return None
        try:
            thickness = float(thick_raw)
        except ValueError:
            messagebox.showwarning("Проверка", "Толщина должна быть числом. Пример: 0.36")
            return None
        if not score_str:
            messagebox.showwarning("Проверка", "Выберите оценку")
            return None
        score = int(score_str.split()[0])

        return {
            'sidak_num': sidak,
            'supplier_name': supplier,
            'material': material,
            'glue': glue,
            'thickness': thickness,
            'score': score,
        }

    def add_film():
        film = validate_and_collect()
        if not film:
            return
        films_data.append(film)
        # в таблицу
        tv.insert('', 'end', values=(film['sidak_num'], film['supplier_name'], film['thickness'], film['score']))
        # очистка
        combo_sidak.set('')
        entry_supplier.delete(0, ctk.END)
        entry_thickness.delete(0, ctk.END)
        combo_sidak.focus_set()

    def delete_selected():
        sel = tv.selection()
        if not sel:
            return
        idxs = [tv.index(i) for i in sel]
        for i in sorted(idxs, reverse=True):
            del films_data[i]
        for i in sel:
            tv.delete(i)

    def make_protocol_memory():
        nonlocal wb, ws, current_film_index
        global protocol_title
        if not films_data:
            messagebox.showwarning("Ошибка", "Сначала добавьте хотя бы одну плёнку")
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Лист1"
        # Сохраняем заголовок в глобальную переменную, чтобы его можно было использовать при сохранении файла
        protocol_title = build_protocol(
            ws,
            films_data,
            test_type=combo_test_type.get(),
            author=combo_author.get(),
            participant=combo_participant.get()
        )
        # включаем панель изображений
        current_film_index = 0
        update_image_panel_state(enabled=True)
        update_current_film_label()
        btn_save.configure(state='normal')
        messagebox.showinfo("Готово", "Протокол сформирован в памяти. Теперь можно добавить картинки и сохранить файл.")

    # ---- Панель изображений (встроенная) ----
    def find_target_row_by_film(sidak_num):
        for row in range(1, ws.max_row + 1):
            cell_value = ws[f'A{row}'].value
            if cell_value and sidak_num in str(cell_value) and "09:30-17:30" in str(cell_value):
                return row
        return None

    def update_current_film_label():
        if ws is None or not films_data:
            lbl_current_film.configure(text="Нет данных")
            return
        if current_film_index >= len(films_data):
            lbl_current_film.configure(text="Все плёнки обработаны")
        else:
            lbl_current_film.configure(text=f"Текущая плёнка: {films_data[current_film_index]['sidak_num']}")

    def update_image_panel_state(enabled: bool):
        state = 'normal' if enabled else 'disabled'
        btn_img_add.configure(state=state)
        btn_img_skip.configure(state=state)
        btn_img_skip_all.configure(state=state)

    def image_add_one():
        nonlocal current_film_index
        if ws is None:
            messagebox.showwarning("Внимание", "Сначала создайте протокол")
            return
        if current_film_index >= len(films_data):
            messagebox.showinfo("Инфо", "Все плёнки уже обработаны")
            update_image_panel_state(False)
            return
        film = films_data[current_film_index]
        target_row = find_target_row_by_film(film['sidak_num'])
        if not target_row:
            print(f"⚠️ Не найдена строка для пленки {film['sidak_num']}")
            current_film_index += 1
            update_current_film_label()
            return

        file_path = filedialog.askopenfilename(
            title=f"Выберите изображение для {film['sidak_num']}",
            filetypes=[("Image files", "*.jpg *.jpeg *.png *.bmp *.gif")]
        )
        if not file_path:
            return
        try:
            from PIL import Image as PILImage
            pil_img = PILImage.open(file_path)
            max_width, max_height = 220, 400
            orig_width, orig_height = pil_img.size
            width_ratio = max_width / orig_width
            height_ratio = max_height / orig_height
            scale_ratio = min(width_ratio, height_ratio)
            new_width = int(orig_width * scale_ratio)
            new_height = int(orig_height * scale_ratio)

            img = Image(file_path)
            img.width = new_width
            img.height = new_height
            cell_address = f"C{target_row}"
            ws.add_image(img, cell_address)
            print(f"✅ Картинка добавлена для {film['sidak_num']} в ячейку {cell_address}")
        except Exception as e:
            print(f"⚠️ Ошибка при добавлении картинки: {e}")
        finally:
            current_film_index += 1
            if current_film_index >= len(films_data):
                messagebox.showinfo("Завершено", "Все картинки добавлены!")
                update_image_panel_state(False)
            update_current_film_label()

    def image_skip_one():
        nonlocal current_film_index
        current_film_index += 1
        if current_film_index >= len(films_data):
            messagebox.showinfo("Завершено", "Обработка завершена!")
            update_image_panel_state(False)
        update_current_film_label()

    def image_skip_all():
        nonlocal current_film_index
        current_film_index = len(films_data)
        update_current_film_label()
        update_image_panel_state(False)
        messagebox.showinfo("Завершено", "Добавление картинок пропущено!")

    def save_protocol_to_file():
        global protocol_title
        if ws is None:
            messagebox.showwarning("Внимание", "Сначала создайте протокол")
            return
        # очищаем название протокола для использования в имени файла
        safe_title = protocol_title if protocol_title else "Новый_протокол_испытаний"
        # заменяем недопустимые для имени файла символы
        invalid = r'<>:"/\\|?*'
        safe_title = "".join([c if c not in invalid else "_" for c in safe_title])
        # убираем переносы строк
        safe_title = safe_title.replace("\\n", " ").replace("\n", " ").strip()
        # ограничим длину, чтобы не было проблем с ОС
        if len(safe_title) > 120:
            safe_title = safe_title[:120]

        filename = filedialog.asksaveasfilename(
            title="Сохранить протокол как",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"{safe_title}.xlsx"
        )
        if not filename:
            return
        try:
            wb.save(filename)
            messagebox.showinfo("Сохранено", f"Файл сохранён: {filename}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить файл: {e}")



    # --- UI ---
    ctk.set_appearance_mode("light")
    ctk.set_default_color_theme("blue")

    root = ctk.CTk()
    root.title("Протокол локального нагрева — ввод данных")

    # --- Новые выпадающие списки ---
    top = ctk.CTkFrame(root, corner_radius=10)
    top.grid(row=0, column=0, columnspan=3, sticky="ew", padx=8, pady=(8,0))
    top.grid_columnconfigure(1, weight=1)

    ctk.CTkLabel(top, text="Вид испытания").grid(row=0, column=0, sticky="w", padx=10, pady=5)
    combo_test_type = ctk.CTkComboBox(top, values=[
        "Нагрев",
        "Воздействие кофе",
        "Воздействие масла",
        "Воздействие ацетона",
    ])
    combo_test_type.set("Нагрев")
    combo_test_type.grid(row=0, column=1, padx=5, pady=2, sticky="ew")

    ctk.CTkLabel(top, text="Автор").grid(row=1, column=0, sticky="w", padx=10, pady=5)
    combo_author = ctk.CTkComboBox(top, values=[
        "Руководитель службы качества Камынин В.А.",
        "Специалист по качеству Павлова Н.А.",
        "Специалист по качеству Сидорова А.И.",
    ])
    combo_author.set("Руководитель службы качества Камынин В.А.")
    combo_author.grid(row=1, column=1, padx=5, pady=2, sticky="ew")

    ctk.CTkLabel(top, text="Участник").grid(row=2, column=0, sticky="w", padx=10, pady=5)
    combo_participant = ctk.CTkComboBox(top, values=[
        "Ведущий инженер-технолог Мкртчян А.Р.",
        ""
    ])
    combo_participant.set("Ведущий инженер-технолог Мкртчян А.Р.")
    combo_participant.grid(row=2, column=1, padx=5, pady=2, sticky="ew")

    # Левая панель: ввод одной плёнки
    left = ctk.CTkFrame(root, corner_radius=10)
    left.grid(row=1, column=0, sticky="nsew", padx=8, pady=8)
    left.grid_columnconfigure(1, weight=1)

    import sqlite3

    # Загружаем данные из базы
    def load_films():
        conn = sqlite3.connect("films.db")
        cur = conn.cursor()
        cur.execute("SELECT sidak_num, supplier_name, thickness FROM films ORDER BY sidak_num")
        rows = cur.fetchall()
        conn.close()
        return rows

    films = load_films()
    all_sidak = [r[0] for r in films]
    sidak_value = ctk.StringVar(value="")
    ctk.CTkLabel(left, text="Обозначение Sidak").grid(row=0, column=0, sticky="w", padx=10, pady=5)
    def on_select_sidak(choice):
        sidak = choice.strip()
        conn = sqlite3.connect("films.db")
        cur = conn.cursor()
        cur.execute("SELECT supplier_name, thickness FROM films WHERE sidak_num = ?", (sidak,))
        details = cur.fetchone()
        conn.close()

        if details:
            entry_supplier.delete(0, ctk.END)
            entry_supplier.insert(0, details[0])
            entry_thickness.delete(0, ctk.END)
            entry_thickness.insert(0, str(details[1]))
        else:
            entry_supplier.delete(0, ctk.END)
            entry_thickness.delete(0, ctk.END)
    # Вместо CTkComboBox используем CTkEntry для ввода
    combo_sidak = ctk.CTkEntry(left)
    combo_sidak.grid(row=0, column=1, sticky="ew", padx=5, pady=5)
    
    # Создаем скроллируемую рамку для выпадающего списка, но пока не размещаем ее
    results_frame = ctk.CTkScrollableFrame(root, width=combo_sidak._current_width, height=200)
    
    # Глобальная переменная для отслеживания открытого состояния
    global is_dropdown_open
    is_dropdown_open = False
    
    # Глобальная переменная для хранения данных
    global all_sidak_data
    all_sidak_data = []

    def select_from_list(sidak_num):
        global is_dropdown_open
        combo_sidak.delete(0, 'end')
        combo_sidak.insert(0, sidak_num)
        
        # Очищаем и скрываем выпадающий список
        results_frame.place_forget()
        is_dropdown_open = False

        # Заполняем остальные поля, как это делала on_select_sidak
        try:
            conn = sqlite3.connect("films.db")
            cur = conn.cursor()
            cur.execute("SELECT supplier_name, thickness FROM films WHERE sidak_num = ?", (sidak_num,))
            details = cur.fetchone()
            conn.close()
            if details:
                entry_supplier.delete(0, ctk.END)
                entry_supplier.insert(0, details[0])
                entry_thickness.delete(0, ctk.END)
                entry_thickness.insert(0, str(details[1]))
            else:
                entry_supplier.delete(0, ctk.END)
                entry_thickness.delete(0, ctk.END)
        except sqlite3.Error as e:
            messagebox.showerror("Ошибка", f"Не удалось подключиться к базе данных: {e}")
            return

    def filter_sidak(event=None):
        global is_dropdown_open
        text = combo_sidak.get().strip().lower()

        # Очищаем выпадающий список
        for widget in results_frame.winfo_children():
            widget.destroy()

        # Фильтруем данные из базы
        try:
            conn = sqlite3.connect("films.db")
            cur = conn.cursor()
            cur.execute("SELECT sidak_num FROM films WHERE lower(sidak_num) LIKE ? ORDER BY sidak_num", (f'%{text}%',))
            filtered_sidak = [r[0] for r in cur.fetchall()]
            conn.close()
        except sqlite3.Error as e:
            messagebox.showerror("Ошибка", f"Не удалось подключиться к базе данных: {e}")
            return

        if not filtered_sidak:
            results_frame.place_forget()
            is_dropdown_open = False
            return

        # Получаем координаты поля ввода
        x_pos = combo_sidak.winfo_rootx() - root.winfo_rootx()
        y_pos = combo_sidak.winfo_rooty() - root.winfo_rooty() + combo_sidak.winfo_height()

        # Размещаем выпадающий список
        results_frame.place(x=x_pos, y=y_pos)
        is_dropdown_open = True
        
        for sidak in filtered_sidak:
            btn = ctk.CTkButton(results_frame, text=sidak, command=lambda s=sidak: select_from_list(s))
            btn.pack(fill="x", pady=2)

    # Привязываем событие ввода к функции фильтрации
    combo_sidak.bind("<KeyRelease>", filter_sidak)

    ctk.CTkLabel(left, text="Поставщик").grid(row=1, column=0, sticky="w", padx=10, pady=5)
    entry_supplier = ctk.CTkEntry(left)
    entry_supplier.grid(row=1, column=1, sticky="ew", padx=5, pady=5)

    ctk.CTkLabel(left, text="Листовой материал").grid(row=2, column=0, sticky="w", padx=10, pady=5)
    combo_material = ctk.CTkComboBox(left, values=[
        "МДФ белый 19 мм  Кроношпан",
        "МДФ 16 мм  Кроношпан",
        "МДФ 19 мм  Egger",
    ])
    combo_material.set("МДФ белый 19 мм  Кроношпан")
    combo_material.grid(row=2, column=1, sticky="ew", padx=5, pady=5)

    ctk.CTkLabel(left, text="Клей").grid(row=3, column=0, sticky="w", padx=10, pady=5)
    combo_glue = ctk.CTkComboBox(left, values=[
        "Клей Perfotak 154/3",
        "Клей Jowat 688.60",
        "Клей Henkel 701.20",
    ])
    combo_glue.set("Клей Perfotak 154/3")
    combo_glue.grid(row=3, column=1, sticky="ew", padx=5, pady=5)

    ctk.CTkLabel(left, text="Толщина").grid(row=4, column=0, sticky="w", padx=10, pady=5)
    entry_thickness = ctk.CTkEntry(left)
    entry_thickness.grid(row=4, column=1, sticky="ew", padx=5, pady=5)

    
    
    ctk.CTkLabel(left, text="Результат (1-5 баллов)").grid(row=5, column=0, sticky="w", padx=10, pady=5)
    combo_score = ctk.CTkComboBox(left, values=["5 баллов", "4 балла", "3 балла", "2 балла", "1 балл"])
    combo_score.set("5 баллов")
    combo_score.grid(row=5, column=1, sticky="ew", padx=5, pady=5)

    btn_add = ctk.CTkButton(left, text="Добавить в список", command=add_film)
    btn_add.grid(row=6, column=0, columnspan=2, sticky="ew", padx=5, pady=(20, 5))

    btn_delete = ctk.CTkButton(left, text="Удалить выбранную", command=delete_selected, fg_color="red", hover_color="darkred")
    btn_delete.grid(row=7, column=0, columnspan=2, sticky="ew", padx=5, pady=5)

    # Правая панель: таблица
    right = ctk.CTkFrame(root, corner_radius=10)
    right.grid(row=1, column=1, sticky="nsew", padx=8, pady=8, rowspan=3)
    right.grid_columnconfigure(0, weight=1)

    tv_frame = ctk.CTkFrame(right)
    tv_frame.pack(fill="both", expand=True)

    tv = ttk.Treeview(tv_frame, columns=('sidak', 'supplier', 'thickness', 'score'), show='headings')
    tv.heading('sidak', text='Sidak')
    tv.heading('supplier', text='Поставщик')
    tv.heading('thickness', text='Толщина')
    tv.heading('score', text='Оценка')

    tv.column('sidak', width=100)
    tv.column('supplier', width=150)
    tv.column('thickness', width=80)
    tv.column('score', width=100)
    tv.pack(fill="both", expand=True, padx=5, pady=5)
    
    # ---- Панель управления (встроена в главный UI) ----
    control_panel = ctk.CTkFrame(root, corner_radius=10)
    control_panel.grid(row=2, column=0, sticky="ew", padx=8, pady=8)
    control_panel.grid_columnconfigure(0, weight=1)

    btn_create_protocol = ctk.CTkButton(control_panel, text="Сформировать протокол", command=make_protocol_memory)
    btn_create_protocol.grid(row=0, column=0, sticky="ew", padx=5, pady=5)

    # ---- Панель для картинок (встроена) ----
    img_frame = ctk.CTkFrame(root, corner_radius=10)
    img_frame.grid(row=3, column=0, sticky="ew", padx=8, pady=8)
    img_frame.grid_columnconfigure(0, weight=1)
    
    lbl_current_film = ctk.CTkLabel(img_frame, text="Нет данных", font=("Arial", 10, "bold"))
    lbl_current_film.grid(row=0, column=0, columnspan=3, pady=5, padx=5)

    btn_img_add = ctk.CTkButton(img_frame, text="Добавить картинку", command=image_add_one, fg_color="green", hover_color="darkgreen")
    btn_img_add.grid(row=1, column=0, padx=5, pady=5, sticky="ew")

    btn_img_skip = ctk.CTkButton(img_frame, text="Пропустить эту", command=image_skip_one, fg_color="yellow", hover_color="darkyellow", text_color="black")
    btn_img_skip.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

    btn_img_skip_all = ctk.CTkButton(img_frame, text="Пропустить все", command=image_skip_all, fg_color="red", hover_color="darkred")
    btn_img_skip_all.grid(row=1, column=2, padx=5, pady=5, sticky="ew")
    
    btn_save = ctk.CTkButton(img_frame, text="Сохранить файл", command=save_protocol_to_file, state="disabled")
    btn_save.grid(row=2, column=0, columnspan=3, padx=5, pady=5, sticky="ew")

    update_image_panel_state(False)

    root.grid_columnconfigure(1, weight=1)
    root.grid_rowconfigure(1, weight=1)

    combo_sidak.focus_set()
    root.mainloop()


# ============================
# Совместимость с прежним интерфейсом (парсер текста)
# ============================

def parse_input_text(text):
    blocks = text.split('Заполните параметры по пленкам')[1:]
    films_data = []

    for block in blocks:
        film = {}
        lines = block.strip().split('\n')

        for line in lines:
            line = line.strip()
            if not line:
                continue

            if 'Обозначение Sidak - ' in line:
                film['sidak_num'] = line.split(' - ')[1].strip()
            elif 'Наименование у поставщика - ' in line:
                film['supplier_name'] = line.split(' - ')[1].strip()
            elif 'Введите толщину пленки' in line:
                thickness_str = line.split(' - ')[1].strip()
                film['thickness'] = float(thickness_str.replace(',', '.'))
            elif 'Результат испытания по 5 бальной шкале - ' in line:
                score_str = line.split(' - ')[1].strip()
                film['score'] = int(score_str.split(' ')[0])

        if film:
            films_data.append(film)

    return films_data


# ============================
# Точка входа
# ============================
if __name__ == "__main__":
    try:
        show_input_form()
    except Exception as e:
        messagebox.showerror("Критическая ошибка", f"Произошла непредвиденная ошибка: {e}\n\n{traceback.format_exc()}")